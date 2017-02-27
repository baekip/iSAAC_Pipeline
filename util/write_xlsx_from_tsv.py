#!/usr/bin/python

import sys
import os
import argparse

#ind_python_library="/BiO/BioTools/pylib/lib/python2.7/site-packages"
#if os.path.exists(ind_python_library):
#    sys.path.append(ind_python_library)
import openpyxl as px

def xlsx_read(input_file, output_file):
    wb = px.Workbook()
    ws1 = wb.active
    ws1.title = "Annotation"
    ws2 = wb.create_sheet(title="Appendix_annotation")
    ws3 = wb.create_sheet(title="Appendix_region")

    region_descript_file = '/home/shsong/work/Pipeline/dnaseq/descript/snpeff.anno.txt'
    annotation_descript_file = '/home/shsong/work/Pipeline/dnaseq/descript/annotation_descript.txt'

    with open(input_file) as input_fp:
        for line in input_fp:
            units = line.strip().split('\t')
            list_output = []
            for idx in range(len(units)):
                col_name = str(units[idx])
                list_output.append(col_name)
            ws1.append(list_output)

    with open(annotation_descript_file) as input_fp:
        for line in input_fp:
            units = line.strip().split('\t')
            list_output = []
            for idx in range(len(units)):
                col_name = str(units[idx])
                list_output.append(col_name)
            ws2.append(list_output)

    with open(region_descript_file) as input_fp:
        for line in input_fp:
            units = line.strip().split('\t')
            list_output = []
            for idx in range(len(units)):
                col_name = str(units[idx])
                list_output.append(col_name)
            ws3.append(list_output)

    wb.save("%s" %output_file)
    return

def usage():
    message='''
python %s

-i, --input      : excel xlsx file

##optional
-o, --output : default(output.xlsx)
''' %sys.argv[0]
    print message


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input')
    parser.add_argument('-o', '--output', default="output.xlsx")
    parser.add_argument('-v', dest='verbose', action='store_true')
    args = parser.parse_args()
    try:
        len(args.input) > 0

    except:
        usage()
        sys.exit(2)

    xlsx_read(args.input, args.output)

if __name__ == '__main__':
    main()
